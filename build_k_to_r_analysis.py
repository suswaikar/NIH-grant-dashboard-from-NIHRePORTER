#!/usr/bin/env python
"""
K-to-R Transition Analysis
===========================
Reads the DoM K award tracking spreadsheet, queries NIH RePORTER for
subsequent (non-K) grants received AFTER each investigator's K support
period, and produces a summary Excel workbook.

Usage:
    python build_k_to_r_analysis.py

Outputs:
    K_awardees_summary_2016_2026.xlsx  (in the Data Requests folder)
        Tab 1: K Awardees 2016-2026       — one row per awardee
        Tab 2: NIH Grants - BU_BMC        — post-K non-K grants at BU/BMC
        Tab 3: NIH Grants - External      — post-K non-K grants elsewhere
        Tab 4: Cross-Reference            — discrepancies between our list and RePORTER

Requirements:
    pip install pandas openpyxl requests
"""

import pandas as pd
import re
import requests
import time
import sys
import io
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

# ── PATHS ────────────────────────────────────────────────────────────────────
BASE = Path(
    r"C:\Users\swaikar\OneDrive - Boston University\Department of Medicine"
    r"\Chair advisory committee for Evans\Data Requests"
)
SRC_FILE = BASE / "K award 2016 - 2026.xlsx"
OUT_FILE = BASE / "K_awardees_summary_2016_2026.xlsx"

API_URL = "https://api.reporter.nih.gov/v2/projects/search"

BU_ORGS = frozenset([
    "BOSTON UNIVERSITY",
    "BOSTON MEDICAL CENTER",
    "BOSTON UNIVERSITY MEDICAL CAMPUS",
])

# K-mechanism activity codes (to exclude from "subsequent" grants)
K_CODES = frozenset([
    "K01", "K02", "K05", "K06", "K07", "K08", "K11", "K12",
    "K22", "K23", "K24", "K25", "K26", "K43", "K76", "K99", "K00",
    "KL1", "KL2",
])

# ── SECTION NAME STANDARDIZATION ────────────────────────────────────────────
SECTION_MAP = {
    "Cardiology": "Cardiovascular Medicine",
    "Cardiovascula Center": "Cardiovascular Medicine",
    "Cardiovascular Medicine": "Cardiovascular Medicine",
    "Clin Epi": "Clinical Epidemiology",
    "Clinical Epidemiology": "Clinical Epidemiology",
    "Endocrine": "Endocrinology",
    "Endocrinology": "Endocrinology",
    "GI": "Gastroenterology",
    "Gastroenterology": "Gastroenterology",
    "GIM": "General Internal Medicine",
    "General Internal Medicine": "General Internal Medicine",
    "Geriatrics": "Geriatrics",
    "geriatrics": "Geriatrics",
    "Hem & Medical Onc": "Hematology/Oncology",
    "Hematology/Oncology": "Hematology/Oncology",
    "Hem/Onc": "Hematology/Oncology",
    "Infectious Disease": "Infectious Diseases",
    "Infectious Diseases": "Infectious Diseases",
    "Infectous Disease": "Infectious Diseases",
    "Nephrology": "Nephrology",
    "Renal": "Nephrology",
    "Pulmonary": "Pulmonary",
    "Pulmonary Center": "Pulmonary",
    "Rheumatology": "Rheumatology",
    "Vascular Biology": "Vascular Biology",
}

# ── NAME NORMALIZATION ───────────────────────────────────────────────────────
NAME_MAP = {
    "Elliot Hagedorn": "Elliott Hagedorn",
    "Titi Ilori": "Titilayo Ilori",
    "Andrew BERICAL": "Andrew Berical",
    "Kostas Alysandratos": "Konstantinos Alysandratos",
}


def clean_name(name: str) -> str:
    """Strip degrees, normalize Last,First → First Last, remove middle initials."""
    name = re.sub(
        r",?\s*(MD|PhD|DO|MPH|MS|MA|DrPH|ScD)\b\.?", "", str(name), flags=re.IGNORECASE
    )
    name = re.sub(r"\s+", " ", name).strip()
    if "," in name:
        parts = [p.strip() for p in name.split(",", 1)]
        if len(parts) == 2 and parts[1]:
            name = parts[1] + " " + parts[0]
    # Remove standalone middle initials (single uppercase letter)
    name = re.sub(r"\s+[A-Z]\.?\s+", " ", name)
    name = re.sub(r"\s+", " ", name).strip()
    # Title-case all-caps words (but leave mixed-case alone)
    name = " ".join(
        w.capitalize() if w.isupper() and len(w) > 1 else w for w in name.split()
    )
    return NAME_MAP.get(name, name)


# ── STEP 1: READ K AWARD SOURCE FILE ────────────────────────────────────────
def load_k_awardees() -> pd.DataFrame:
    """Parse all FY sheets from the K award tracking spreadsheet."""
    xls = pd.ExcelFile(SRC_FILE)
    all_rows = []

    for sheet in xls.sheet_names:
        raw = pd.read_excel(SRC_FILE, sheet_name=sheet, header=None)
        # Find header row (contains "Section")
        header_row = None
        for r in range(min(5, len(raw))):
            vals = [str(v).strip().lower() for v in raw.iloc[r] if pd.notna(v)]
            if "section" in vals:
                header_row = r
                break
        if header_row is None:
            continue

        df = pd.read_excel(SRC_FILE, sheet_name=sheet, header=header_row)
        if "Section" not in df.columns or "Name" not in df.columns:
            continue

        df = df[df["Section"].notna() & df["Name"].notna()].copy()
        df["Section"] = (
            df["Section"].str.strip().map(SECTION_MAP).fillna(df["Section"].str.strip())
        )
        df["Name"] = df["Name"].astype(str).apply(clean_name)
        df["FY"] = sheet
        # Extract FY number for comparisons
        fy_match = re.search(r"\d{4}", sheet)
        df["FY_Num"] = int(fy_match.group()) if fy_match else 0

        # 50% salary gap
        gap_col = [
            c
            for c in df.columns
            if "50" in str(c) and ("salary" in str(c).lower() or "gap" in str(c).lower())
        ]
        df["SalaryGap"] = (
            pd.to_numeric(df[gap_col[0]], errors="coerce") if gap_col else float("nan")
        )

        all_rows.append(df[["Name", "Section", "FY", "FY_Num", "SalaryGap"]])

    return pd.concat(all_rows, ignore_index=True)


# ── STEP 2: QUERY NIH REPORTER ──────────────────────────────────────────────
def fetch_paginated(criteria: dict, limit: int = 500) -> list:
    """Fetch all records matching criteria, paginating in chunks."""
    results = []
    offset = 0
    while True:
        payload = {
            "criteria": criteria,
            "include_fields": [
                "ProjectNum", "ProjectTitle", "ContactPiName",
                "PrincipalInvestigators", "Organization", "FiscalYear",
                "AwardAmount", "ActivityCode", "CoreProjectNum",
                "AgencyIcAdmin", "ApplId", "IsActive",
                "ProjectStartDate", "ProjectEndDate",
                "DirectCostAmt", "IndirectCostAmt",
            ],
            "offset": offset,
            "limit": limit,
            "sort_field": "fiscal_year",
            "sort_order": "desc",
        }
        r = requests.post(API_URL, json=payload, timeout=30)
        r.raise_for_status()
        data = r.json()
        results.extend(data["results"])
        total = data["meta"]["total"]
        offset += limit
        if offset >= total or offset >= 5000:
            break
        time.sleep(0.15)
    return results


def query_nih_for_investigators(names: list[str]) -> pd.DataFrame:
    """Query NIH RePORTER for all grants associated with each investigator."""
    all_grants = []

    for i, name in enumerate(names):
        parts = name.split()
        first, last = parts[0], parts[-1]
        results = []

        for yr_range in [list(range(2015, 2021)), list(range(2021, 2027))]:
            try:
                chunk = fetch_paginated({
                    "pi_names": [{"first_name": first, "last_name": last}],
                    "fiscal_years": yr_range,
                })
                results.extend(chunk)
            except Exception as e:
                print(f"  WARN: {first} {last} ({yr_range[0]}-{yr_range[-1]}): {e}")
            time.sleep(0.2)

        # Filter: PI must actually match (not just partial name collision)
        count = 0
        for g in results:
            pis = g.get("principal_investigators") or []
            pi_match = any(
                last.upper() in (p.get("last_name", "") or "").upper()
                and first.upper()[:3] in (p.get("first_name", "") or "").upper()[:3]
                for p in pis
            )
            if not pi_match:
                continue

            is_contact = any(
                last.upper() in (p.get("last_name", "") or "").upper()
                and first.upper()[:3] in (p.get("first_name", "") or "").upper()[:3]
                and p.get("is_contact_pi", False)
                for p in pis
            )

            org = (g.get("organization") or {}).get("org_name", "")
            ic = g.get("agency_ic_admin") or {}
            ic_abbr = ic.get("abbreviation", "") if isinstance(ic, dict) else str(ic)

            all_grants.append({
                "Name": name,
                "Project Number": g.get("project_num", ""),
                "Core Project": g.get("core_project_num", ""),
                "Activity Code": g.get("activity_code", ""),
                "Title": (g.get("project_title") or "")[:150],
                "Organization": org,
                "Fiscal Year": g.get("fiscal_year"),
                "Award Amount": g.get("award_amount"),
                "Direct Cost": g.get("direct_cost_amt"),
                "Indirect Cost": g.get("indirect_cost_amt"),
                "IC": ic_abbr,
                "Start Date": (g.get("project_start_date") or "")[:10],
                "End Date": (g.get("project_end_date") or "")[:10],
                "Is Active": g.get("is_active"),
                "Contact PI": g.get("contact_pi_name", ""),
                "Is Contact PI": is_contact,
            })
            count += 1

        print(f"  [{i+1}/{len(names)}] {name}: {count} grant-years")

    df = pd.DataFrame(all_grants)
    if len(df):
        df = df.drop_duplicates(subset=["Name", "Project Number", "Fiscal Year"])
    return df


# ── STEP 3: FILTER FALSE POSITIVES ──────────────────────────────────────────
def filter_false_positives(grants: pd.DataFrame) -> pd.DataFrame:
    """Remove grants that are clearly a different person with the same name.

    All grants reaching this step already passed the PI name-match in Step 2
    (last name + first-3 chars of first name in principal_investigators list).
    Step 3 only removes hand-curated same-name collisions; it does NOT
    additionally require BU/BMC org or contact-PI status, since that would drop
    legitimate non-contact co-PI grants at external institutions (e.g., a BU
    awardee listed as co-PI on a BWH/Stanford U01)."""
    filtered = grants.copy()

    # --- Hand-curated same-name collision rules ---
    # Sun Lee: keep only BU/BMC grants or where contact PI matches "LEE-MARQUEZ" / "LEE, SUN Y"
    sun_bad = (filtered["Name"] == "Sun Lee") & ~(
        filtered["Contact PI"].str.upper().str.contains(
            "LEE-MARQUEZ|LEE, SUN Y|LEE,SUN", na=False, regex=True
        )
        | filtered["Organization"].str.upper().isin(BU_ORGS)
    )
    # Sudhir Kumar at Temple/Iowa State — different person
    kumar_bad = (filtered["Name"] == "Sudhir Kumar") & ~filtered[
        "Organization"
    ].str.upper().isin(BU_ORGS)

    removed = sun_bad.sum() + kumar_bad.sum()
    if removed:
        print(f"  Removed {removed} same-name-collision rows (Sun Lee: {sun_bad.sum()}, Sudhir Kumar: {kumar_bad.sum()})")

    filtered = filtered[~sun_bad & ~kumar_bad].copy()
    filtered["Location"] = filtered["Organization"].str.upper().apply(
        lambda x: "BU/BMC" if x in BU_ORGS else "External"
    )
    return filtered


# ── STEP 4: CROSS-REFERENCE ─────────────────────────────────────────────────
def cross_reference(k_awardees: pd.DataFrame, grants: pd.DataFrame) -> pd.DataFrame:
    """Compare our K awardee list against NIH RePORTER K grants.

    Flags:
    - Awardees in our list but NOT found in RePORTER with a K grant
    - K grants in RePORTER at BU/BMC not in our tracking spreadsheet
    - Year mismatches between our list and RePORTER
    """
    issues = []

    # Our awardees grouped
    our = k_awardees.groupby("Name").agg(
        FYs=("FY", lambda x: sorted(x.unique())),
        FY_Nums=("FY_Num", lambda x: sorted(x.unique())),
        Section=("Section", "first"),
    ).reset_index()

    # RePORTER K grants at BU/BMC
    k_grants_rep = grants[
        (grants["Activity Code"].isin(K_CODES))
        & (grants["Organization"].str.upper().isin(BU_ORGS))
    ].copy()

    for _, row in our.iterrows():
        name = row["Name"]
        rep_k = k_grants_rep[k_grants_rep["Name"] == name]
        if len(rep_k) == 0:
            issues.append({
                "Name": name,
                "Section": row["Section"],
                "Issue": "In our list but NO K grant found in NIH RePORTER at BU/BMC",
                "Our Years": ", ".join(row["FYs"]),
                "RePORTER Years": "",
                "RePORTER Grants": "",
            })
        else:
            our_fy = set(row["FY_Nums"])
            rep_fy = set(rep_k["Fiscal Year"].dropna().astype(int))
            our_only = our_fy - rep_fy
            rep_only = rep_fy - our_fy
            if our_only or rep_only:
                issues.append({
                    "Name": name,
                    "Section": row["Section"],
                    "Issue": f"Year mismatch — our list only: {sorted(our_only) if our_only else 'none'}, "
                             f"RePORTER only: {sorted(rep_only) if rep_only else 'none'}",
                    "Our Years": ", ".join(row["FYs"]),
                    "RePORTER Years": ", ".join(str(y) for y in sorted(rep_fy)),
                    "RePORTER Grants": ", ".join(rep_k["Project Number"].unique()[:5]),
                })

    # Check for K grants at BU/BMC for names NOT in our list
    rep_names = set(k_grants_rep["Name"].unique())
    our_names = set(our["Name"].unique())
    # (This won't catch new names since we only queried our list, but it's a placeholder)

    return pd.DataFrame(issues)


# ── STEP 5: BUILD EXCEL ─────────────────────────────────────────────────────
def build_excel(
    summary: pd.DataFrame,
    bu_grants: pd.DataFrame,
    ext_grants: pd.DataFrame,
    xref: pd.DataFrame,
):
    """Write the multi-tab Excel workbook."""
    wb = Workbook()

    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    data_font = Font(name="Arial", size=11)
    money_fmt = '"$"#,##0'
    thin_border = Border(bottom=Side(style="thin", color="D9D9D9"))
    alt_fill = PatternFill("solid", fgColor="F2F7FB")
    warn_fill = PatternFill("solid", fgColor="FFF3CD")

    def write_sheet(ws, headers, data, money_cols=None):
        money_cols = money_cols or []
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
        for i, (_, row) in enumerate(data.iterrows(), 2):
            for col, h in enumerate(headers, 1):
                val = row.get(h, "")
                cell = ws.cell(row=i, column=col, value=val if pd.notna(val) else "")
                cell.font = data_font
                cell.border = thin_border
                if i % 2 == 0:
                    cell.fill = alt_fill
                if h in money_cols:
                    cell.number_format = money_fmt
                    cell.alignment = Alignment(horizontal="right", vertical="center")
        ws.auto_filter.ref = f"A1:{chr(64+len(headers))}{len(data)+1}"
        ws.freeze_panes = "A2"

    # --- Tab 1: Summary ---
    ws1 = wb.active
    ws1.title = "K Awardees 2016-2026"
    headers1 = [
        "Name", "Section", "Years Supported", "Last K Year",
        "Total 50% Salary Gap",
        "Post-K NIH Direct Costs", "Post-K NIH Indirect Costs",
        "Post-K Grant Count",
    ]
    write_sheet(
        ws1, headers1, summary,
        money_cols=["Total 50% Salary Gap", "Post-K NIH Direct Costs", "Post-K NIH Indirect Costs"],
    )
    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 28
    ws1.column_dimensions["C"].width = 55
    ws1.column_dimensions["D"].width = 12
    ws1.column_dimensions["E"].width = 22
    ws1.column_dimensions["F"].width = 24
    ws1.column_dimensions["G"].width = 24
    ws1.column_dimensions["H"].width = 18

    # --- Tab 2: BU/BMC grants ---
    ws2 = wb.create_sheet("NIH Grants - BU_BMC")
    grant_headers = [
        "Name", "Project Number", "Activity Code", "Title", "Organization",
        "Fiscal Year", "Direct Cost", "Indirect Cost", "Award Amount",
        "IC", "Contact PI", "Start Date", "End Date", "Is Active",
    ]
    write_sheet(ws2, grant_headers, bu_grants, money_cols=["Direct Cost", "Indirect Cost", "Award Amount"])
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 24
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 50
    ws2.column_dimensions["E"].width = 35
    for c in "FGHIJKLMN":
        ws2.column_dimensions[c].width = 15

    # --- Tab 3: External grants ---
    ws3 = wb.create_sheet("NIH Grants - External")
    write_sheet(ws3, grant_headers, ext_grants, money_cols=["Direct Cost", "Indirect Cost", "Award Amount"])
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 24
    ws3.column_dimensions["C"].width = 14
    ws3.column_dimensions["D"].width = 50
    ws3.column_dimensions["E"].width = 35
    for c in "FGHIJKLMN":
        ws3.column_dimensions[c].width = 15

    # --- Tab 4: Cross-Reference ---
    ws4 = wb.create_sheet("Cross-Reference")
    xref_headers = ["Name", "Section", "Issue", "Our Years", "RePORTER Years", "RePORTER Grants"]
    write_sheet(ws4, xref_headers, xref)
    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["B"].width = 28
    ws4.column_dimensions["C"].width = 70
    ws4.column_dimensions["D"].width = 45
    ws4.column_dimensions["E"].width = 30
    ws4.column_dimensions["F"].width = 40
    # Highlight issue rows
    for row_idx in range(2, len(xref) + 2):
        for col_idx in range(1, len(xref_headers) + 1):
            ws4.cell(row=row_idx, column=col_idx).fill = warn_fill

    wb.save(OUT_FILE)


# ── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    print("Step 1: Reading K award spreadsheet...")
    k_data = load_k_awardees()
    unique_names = sorted(k_data["Name"].unique())
    print(f"  Found {len(unique_names)} unique K awardees across {k_data['FY'].nunique()} fiscal years\n")

    # Build per-awardee summary
    awardee_summary = k_data.groupby(["Name", "Section"]).agg(
        Years=("FY", lambda x: ", ".join(sorted(x.unique()))),
        LastKYear=("FY_Num", "max"),
        TotalGap=("SalaryGap", "sum"),
    ).reset_index()
    awardee_summary.columns = [
        "Name", "Section", "Years Supported", "Last K Year", "Total 50% Salary Gap",
    ]

    print("Step 2: Querying NIH RePORTER for all investigators...")
    all_grants = query_nih_for_investigators(unique_names)
    print(f"  Raw: {len(all_grants)} grant-year records\n")

    print("Step 3: Filtering false positives...")
    grants = filter_false_positives(all_grants)
    print(f"  Filtered: {len(grants)} grant-year records for {grants['Name'].nunique()} investigators\n")

    print("Step 4: Cross-referencing our list against RePORTER...")
    xref = cross_reference(k_data, grants)
    print(f"  Found {len(xref)} discrepancies\n")

    print("Step 5: Filtering to non-K grants from K-entry onward...")
    # For each awardee, get their FIRST K support year (anchor point)
    first_k = k_data.groupby("Name")["FY_Num"].min().to_dict()

    # Exclude K-mechanism grants
    non_k = grants[~grants["Activity Code"].isin(K_CODES)].copy()
    print(f"  Non-K grants: {len(non_k)} (removed {len(grants) - len(non_k)} K grants)")

    # Keep grants from fiscal years >= first K year (allows concurrent R awards during K)
    non_k["First_K_Year"] = non_k["Name"].map(first_k)
    post_k = non_k[non_k["Fiscal Year"] >= non_k["First_K_Year"]].copy()
    print(f"  Non-K grants from K-entry onward: {len(post_k)} (removed {len(non_k) - len(post_k)} grants predating K support)")

    # Compute post-K totals per investigator
    post_k_summary = post_k.groupby("Name").agg(
        Direct=("Direct Cost", "sum"),
        Indirect=("Indirect Cost", "sum"),
        GrantCount=("Core Project", "nunique"),
    ).reset_index()

    # Merge into summary
    summary = awardee_summary.merge(post_k_summary, on="Name", how="left")
    summary["Direct"] = summary["Direct"].fillna(0)
    summary["Indirect"] = summary["Indirect"].fillna(0)
    summary["GrantCount"] = summary["GrantCount"].fillna(0).astype(int)
    summary.rename(columns={
        "Direct": "Post-K NIH Direct Costs",
        "Indirect": "Post-K NIH Indirect Costs",
        "GrantCount": "Post-K Grant Count",
    }, inplace=True)
    summary = summary.sort_values(["Section", "Name"]).reset_index(drop=True)

    # Split post-K grants by location
    bu_grants = post_k[post_k["Location"] == "BU/BMC"].sort_values(["Name", "Fiscal Year"]).reset_index(drop=True)
    ext_grants = post_k[post_k["Location"] == "External"].sort_values(["Name", "Fiscal Year"]).reset_index(drop=True)

    print(f"\nStep 6: Building Excel...")
    build_excel(summary, bu_grants, ext_grants, xref)
    print(f"  Saved to: {OUT_FILE}")
    print(f"  Summary tab: {len(summary)} awardees")
    print(f"  BU/BMC grants tab: {len(bu_grants)} rows")
    print(f"  External grants tab: {len(ext_grants)} rows")
    print(f"  Cross-reference tab: {len(xref)} issues")

    # Print quick stats
    has_post_k = summary[summary["Post-K Grant Count"] > 0]
    print(f"\n── Summary ──")
    print(f"  K awardees with post-K NIH funding: {len(has_post_k)}/{len(summary)} ({100*len(has_post_k)/len(summary):.0f}%)")
    print(f"  Total post-K direct costs: ${summary['Post-K NIH Direct Costs'].sum():,.0f}")
    print(f"  Total post-K indirect costs: ${summary['Post-K NIH Indirect Costs'].sum():,.0f}")
    print(f"  Total DoM salary gap support: ${summary['Total 50% Salary Gap'].sum():,.0f}")


if __name__ == "__main__":
    main()
