#!/usr/bin/env python
"""
Pilot / Junior Faculty Award ROI Analysis
==========================================
Reads the DoM Award Tracker (Pilot and Junior Faculty awards 2022-2026), queries
NIH RePORTER for each awardee's subsequent NIH funding, and computes ROI
under two definitions:

NOTE: GT97 awards are intentionally excluded — they are an accounting mechanism
allowing well-funded investigators to report less than 100% federal effort,
not a true pilot/research investment. Including them inflated the ROI artifact
(GT97 awards average ~$7k and primarily go to senior faculty with massive
existing NIH portfolios).

    GENEROUS   : all NIH grants active in fiscal year >= first DoM award year
                 (matches Option B from the K-to-R analysis; allows funding
                 that started before the pilot but was active during/after)

    CONSERVATIVE: only NIH grants whose project_start_date year is >= first
                  DoM award year (truly new awards initiated during/after
                  the DoM investment)

Outputs:
    DoM_Pilot_ROI_Summary.xlsx
        Tab 1: Investigator Summary    — one row per awardee, both ROI numbers
        Tab 2: By Award Type           — Pilot vs GT97 vs Junior aggregates
        Tab 3: NIH Grants - BU/BMC     — per-grant detail (generous filter)
        Tab 4: NIH Grants - External   — per-grant detail (generous filter)
        Tab 5: Cross-Reference         — comparison vs user's manual exports

Usage:
    python build_pilot_roi_analysis.py
"""

import pandas as pd
import re
import requests
import time
import sys
import io
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

# ── PATHS ────────────────────────────────────────────────────────────────────
BASE = Path(
    r"C:\Users\swaikar\OneDrive - Boston University\Department of Medicine"
    r"\Chair advisory committee for Evans"
)
TRACKER = BASE / "DoM Award Tracker.xlsx"
USER_EXPORT_ALL = BASE / "Pilot, GT97, Jr Fac_NIH RePORTER 2022-2026.xlsx"
USER_EXPORT_SUB = BASE / "Pilot, Jr Fac_NIH RePORTER 2022-2026.xlsx"
OUT_FILE = BASE / "shared drive_041626" / "Research Advisory Committee" / "Data Requests" / "DoM_Pilot_ROI_Summary.xlsx"
CACHE_FILE = BASE / "NIH awards" / "_pilot_raw_grants_cache.csv"

API_URL = "https://api.reporter.nih.gov/v2/projects/search"

BU_ORGS = frozenset([
    "BOSTON UNIVERSITY",
    "BOSTON MEDICAL CENTER",
    "BOSTON UNIVERSITY MEDICAL CAMPUS",
])

# Internal DoM awards are not in RePORTER; nothing to exclude as a "self-match"
# But we DO want to know what's NEW — so for conservative count we filter by start date.

# ── NAME NORMALIZATION (shared with K analysis) ─────────────────────────────
NAME_MAP = {
    "Elliot Hagedorn": "Elliott Hagedorn",
    "Titi Ilori": "Titilayo Ilori",
    "Andrew BERICAL": "Andrew Berical",
    "Kostas Alysandratos": "Konstantinos Alysandratos",
    "Konstantinos-Dionysios Alysandratos": "Konstantinos Alysandratos",
    "Konstantinos-Dionysios": "Konstantinos Alysandratos",  # tracker has truncated name
}


def clean_name(name: str) -> str:
    name = re.sub(
        r",?\s*(MD|PhD|DO|MPH|MS|MA|DrPH|ScD)\b\.?", "", str(name), flags=re.IGNORECASE
    )
    name = re.sub(r"\s+", " ", name).strip()
    if "," in name:
        parts = [p.strip() for p in name.split(",", 1)]
        if len(parts) == 2 and parts[1]:
            name = parts[1] + " " + parts[0]
    name = re.sub(r"\s+[A-Z]\.?\s+", " ", name)
    name = re.sub(r"\s+", " ", name).strip()
    name = " ".join(
        w.capitalize() if w.isupper() and len(w) > 1 else w for w in name.split()
    )
    return NAME_MAP.get(name, name)


def ay_to_fy(ay: str) -> int:
    """Convert 'AY22' → 2022. Treat AY≈FY per user's call."""
    m = re.search(r"(\d{2,4})", str(ay))
    if not m:
        return 0
    yr = int(m.group(1))
    return 2000 + yr if yr < 100 else yr


# ── STEP 1: LOAD DOM AWARDS ─────────────────────────────────────────────────
def load_dom_awards() -> pd.DataFrame:
    df = pd.read_excel(TRACKER, sheet_name="Combined")
    df["Name"] = df["Name"].astype(str).apply(clean_name)
    df["FY_Num"] = df["FY"].apply(ay_to_fy)
    df["Amount"] = pd.to_numeric(df["Amount"], errors="coerce").fillna(0)
    df["Award"] = df["Award"].astype(str).str.strip()
    # Standardize award labels
    df["Award"] = df["Award"].replace({"Junior Award": "Junior"})
    # Exclude GT97 — accounting mechanism, not a true pilot award (see header docstring)
    n_gt97 = (df["Award"] == "GT97").sum()
    df = df[df["Award"] != "GT97"].copy()
    print(f"  (excluded {n_gt97} GT97 award rows — see header note)")
    return df


# ── STEP 2: QUERY NIH REPORTER ──────────────────────────────────────────────
def fetch_paginated(criteria: dict, limit: int = 500) -> list:
    results = []
    offset = 0
    while True:
        payload = {
            "criteria": criteria,
            "include_fields": [
                "ProjectNum", "ProjectTitle", "ContactPiName",
                "PrincipalInvestigators", "Organization", "FiscalYear",
                "AwardAmount", "ActivityCode", "CoreProjectNum",
                "AgencyIcAdmin", "ApplId", "IsActive",
                "ProjectStartDate", "ProjectEndDate",
                "DirectCostAmt", "IndirectCostAmt",
            ],
            "offset": offset,
            "limit": limit,
            "sort_field": "fiscal_year",
            "sort_order": "desc",
        }
        r = requests.post(API_URL, json=payload, timeout=30)
        r.raise_for_status()
        data = r.json()
        results.extend(data["results"])
        total = data["meta"]["total"]
        offset += limit
        if offset >= total or offset >= 5000:
            break
        time.sleep(0.15)
    return results


def query_nih(names: list[str]) -> pd.DataFrame:
    all_grants = []
    for i, name in enumerate(names):
        parts = name.split()
        if len(parts) < 2:
            print(f"  [{i+1}/{len(names)}] SKIP (single token): {name}")
            continue
        first, last = parts[0], parts[-1]
        results = []
        # Cover FY 2021-2026 (AY22 first awards, but allow one earlier year buffer)
        for yr_range in [list(range(2021, 2027))]:
            try:
                results.extend(fetch_paginated({
                    "pi_names": [{"first_name": first, "last_name": last}],
                    "fiscal_years": yr_range,
                }))
            except Exception as e:
                print(f"  WARN: {first} {last}: {e}")
            time.sleep(0.2)

        count = 0
        for g in results:
            pis = g.get("principal_investigators") or []
            pi_match = any(
                last.upper() in (p.get("last_name", "") or "").upper()
                and first.upper()[:3] in (p.get("first_name", "") or "").upper()[:3]
                for p in pis
            )
            if not pi_match:
                continue
            is_contact = any(
                last.upper() in (p.get("last_name", "") or "").upper()
                and first.upper()[:3] in (p.get("first_name", "") or "").upper()[:3]
                and p.get("is_contact_pi", False)
                for p in pis
            )
            org = (g.get("organization") or {}).get("org_name", "")
            ic = g.get("agency_ic_admin") or {}
            ic_abbr = ic.get("abbreviation", "") if isinstance(ic, dict) else str(ic)

            all_grants.append({
                "Name": name,
                "Project Number": g.get("project_num", ""),
                "Core Project": g.get("core_project_num", ""),
                "Activity Code": g.get("activity_code", ""),
                "Title": (g.get("project_title") or "")[:150],
                "Organization": org,
                "Fiscal Year": g.get("fiscal_year"),
                "Award Amount": g.get("award_amount") or 0,
                "Direct Cost": g.get("direct_cost_amt") or 0,
                "Indirect Cost": g.get("indirect_cost_amt") or 0,
                "IC": ic_abbr,
                "Start Date": (g.get("project_start_date") or "")[:10],
                "End Date": (g.get("project_end_date") or "")[:10],
                "Is Active": g.get("is_active"),
                "Contact PI": g.get("contact_pi_name", ""),
                "Is Contact PI": is_contact,
            })
            count += 1
        print(f"  [{i+1}/{len(names)}] {name}: {count} grant-years")

    df = pd.DataFrame(all_grants)
    if len(df):
        df = df.drop_duplicates(subset=["Name", "Project Number", "Fiscal Year"])
    return df


# ── STEP 3: FILTER FALSE POSITIVES ──────────────────────────────────────────
def filter_grants(grants: pd.DataFrame) -> pd.DataFrame:
    """All grants in `grants` already passed the PI name-match in Step 2 (last name +
    first-3 chars of first name in principal_investigators list). Step 3 only removes
    hand-curated same-name collisions; it does NOT additionally require BU/BMC org
    or contact-PI status, since that would drop legitimate non-contact co-PI grants
    at external institutions."""
    keep = grants.copy()

    # Hand-curated same-name collisions (Sun Lee at non-BU, Sudhir Kumar at Temple/Iowa State)
    sun_bad = (keep["Name"] == "Sun Lee") & ~(
        keep["Contact PI"].str.upper().str.contains(
            "LEE-MARQUEZ|LEE, SUN Y|LEE,SUN", na=False, regex=True
        )
        | keep["Organization"].str.upper().isin(BU_ORGS)
    )
    kumar_bad = (keep["Name"] == "Sudhir Kumar") & ~keep["Organization"].str.upper().isin(BU_ORGS)
    removed = sun_bad.sum() + kumar_bad.sum()
    if removed:
        print(f"  Removed {removed} same-name-collision rows (Sun Lee: {sun_bad.sum()}, Sudhir Kumar: {kumar_bad.sum()})")
    keep = keep[~sun_bad & ~kumar_bad].copy()
    keep["Location"] = keep["Organization"].str.upper().apply(
        lambda x: "BU/BMC" if x in BU_ORGS else "External"
    )
    return keep


# ── STEP 4: COMPUTE ROI (TWO DEFINITIONS) ───────────────────────────────────
def compute_roi(dom: pd.DataFrame, grants: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    # First DoM year per investigator
    first_dom = dom.groupby("Name")["FY_Num"].min().to_dict()

    # GENEROUS: grants active in FY >= first DoM year
    gen = grants[grants["Fiscal Year"].notna()].copy()
    gen["First_DoM_FY"] = gen["Name"].map(first_dom)
    gen = gen[gen["Fiscal Year"] >= gen["First_DoM_FY"]].copy()

    # CONSERVATIVE: project_start_date year >= first DoM year
    grants["Start Year"] = pd.to_datetime(grants["Start Date"], errors="coerce").dt.year
    cons = grants[grants["Start Year"].notna()].copy()
    cons["First_DoM_FY"] = cons["Name"].map(first_dom)
    cons = cons[cons["Start Year"] >= cons["First_DoM_FY"]].copy()

    # Per-investigator DoM totals + award types
    dom_summary = dom.groupby("Name").agg(
        Section=("Section", "first"),
        FirstDoMFY=("FY_Num", "min"),
        LastDoMFY=("FY_Num", "max"),
        DoMTotal=("Amount", "sum"),
        DoMAwards=("Award", lambda x: ", ".join(sorted(set(x)))),
        DoMYears=("FY", lambda x: ", ".join(sorted(set(x)))),
    ).reset_index()

    # Pivot DoM amounts by type
    by_type = dom.pivot_table(
        index="Name", columns="Award", values="Amount", aggfunc="sum", fill_value=0
    ).reset_index()
    for col in ["Pilot", "Junior"]:
        if col not in by_type.columns:
            by_type[col] = 0
    by_type = by_type[["Name", "Pilot", "Junior"]]
    by_type.columns = ["Name", "Pilot $", "Junior $"]
    dom_summary = dom_summary.merge(by_type, on="Name", how="left")

    # NIH funding aggregates per investigator
    def agg_funding(df, label):
        s = df.groupby("Name").agg(
            **{
                f"NIH Direct ({label})": ("Direct Cost", "sum"),
                f"NIH Indirect ({label})": ("Indirect Cost", "sum"),
                f"NIH Total ({label})": ("Award Amount", "sum"),
                f"Grant Count ({label})": ("Core Project", "nunique"),
            }
        ).reset_index()
        return s

    gen_agg = agg_funding(gen, "Generous")
    cons_agg = agg_funding(cons, "Conservative")

    summary = dom_summary.merge(gen_agg, on="Name", how="left").merge(cons_agg, on="Name", how="left")
    for col in summary.columns:
        if col.startswith(("NIH ", "Grant Count")):
            summary[col] = summary[col].fillna(0)

    summary["ROI Generous (Direct/DoM)"] = (
        summary["NIH Direct (Generous)"] / summary["DoMTotal"]
    ).replace([float("inf"), -float("inf")], 0).fillna(0).round(1)
    summary["ROI Conservative (Direct/DoM)"] = (
        summary["NIH Direct (Conservative)"] / summary["DoMTotal"]
    ).replace([float("inf"), -float("inf")], 0).fillna(0).round(1)

    summary = summary.sort_values(["Section", "Name"]).reset_index(drop=True)
    return summary, gen, cons


# ── NARRATIVE: PER-INVESTIGATOR TRAJECTORY ──────────────────────────────────
def build_narrative(dom: pd.DataFrame, gen: pd.DataFrame, cons: pd.DataFrame) -> pd.DataFrame:
    """One row per investigator: DoM history → key NIH grants since.
    Highlights R/U/DP awards (the meaningful ROI signal)."""
    MAJOR_PREFIXES = ("R0", "R1", "R2", "R3", "R5", "U0", "U1", "U2", "U5", "DP", "P0", "P5")

    rows = []
    for name in sorted(dom["Name"].unique()):
        d = dom[dom["Name"] == name].sort_values("FY_Num")
        section = d["Section"].iloc[0]

        # DoM history string
        dom_parts = []
        for _, r in d.iterrows():
            dom_parts.append(f"{r['Award']} {r['FY']} (${r['Amount']:,.0f})")
        dom_str = " + ".join(dom_parts)
        dom_total = d["Amount"].sum()

        # Major grants under each definition (by core project, taking earliest FY)
        def grants_str(df, label):
            sub = df[df["Name"] == name].copy()
            if len(sub) == 0:
                return f"(none {label})"
            # Collapse multi-year grants to one line per core project
            collapsed = sub.sort_values("Fiscal Year").groupby("Core Project").agg(
                ProjectNum=("Project Number", "first"),
                Activity=("Activity Code", "first"),
                IC=("IC", "first"),
                FirstFY=("Fiscal Year", "min"),
                StartDate=("Start Date", "first"),
                Org=("Organization", "first"),
                Direct=("Direct Cost", "sum"),
                Total=("Award Amount", "sum"),
            ).reset_index()
            # Keep major grants only
            major = collapsed[collapsed["Activity"].fillna("").str.startswith(MAJOR_PREFIXES)]
            other = collapsed[~collapsed["Activity"].fillna("").str.startswith(MAJOR_PREFIXES)]
            parts = []
            for _, g in major.sort_values("FirstFY").iterrows():
                org_short = "BU/BMC" if any(o in str(g["Org"]).upper() for o in ["BOSTON UNIVERSITY", "BOSTON MEDICAL"]) else "external"
                parts.append(
                    f"{g['Activity']} {g['ProjectNum']} ({g['IC']}, FY{int(g['FirstFY'])}, ${g['Total']:,.0f}, {org_short})"
                )
            if len(other):
                parts.append(f"+ {len(other)} other (F/T/K/etc.)")
            return "; ".join(parts) if parts else f"(no major awards {label})"

        gen_str = grants_str(gen, "generous")
        cons_str = grants_str(cons, "conservative")

        # Total NIH $ under each
        gen_sub = gen[gen["Name"] == name]
        cons_sub = cons[cons["Name"] == name]
        nih_dir_gen = gen_sub["Direct Cost"].sum()
        nih_dir_cons = cons_sub["Direct Cost"].sum()

        rows.append({
            "Name": name,
            "Section": section,
            "DoM History": dom_str,
            "DoM Total": dom_total,
            "NIH Direct (Generous)": nih_dir_gen,
            "ROI Generous": round(nih_dir_gen / dom_total, 1) if dom_total else 0,
            "Major NIH Grants — Generous (active from first DoM FY)": gen_str,
            "NIH Direct (Conservative)": nih_dir_cons,
            "ROI Conservative": round(nih_dir_cons / dom_total, 1) if dom_total else 0,
            "Major NIH Grants — Conservative (started at/after first DoM FY)": cons_str,
        })
    return pd.DataFrame(rows).sort_values(["Section", "Name"]).reset_index(drop=True)


# ── MISSING-GRANTS INVESTIGATION ────────────────────────────────────────────
def investigate_missing(gen_grants: pd.DataFrame, step3_grants: pd.DataFrame,
                        raw_grants: pd.DataFrame, dom: pd.DataFrame) -> pd.DataFrame:
    """For each grant in user's export but not in our final pipeline, categorize WHY."""
    user_df = pd.read_excel(USER_EXPORT_SUB, sheet_name="Export Sheet")
    user_df["Project Number"] = user_df["Project Number"].astype(str).str.strip()
    gen_proj = set(gen_grants["Project Number"].dropna().astype(str).str.strip())
    step3_proj = set(step3_grants["Project Number"].dropna().astype(str).str.strip())
    raw_proj = set(raw_grants["Project Number"].dropna().astype(str).str.strip())
    awardee_names = set(dom["Name"].unique())
    first_dom = dom.groupby("Name")["FY_Num"].min().to_dict()

    # Map project numbers in raw_grants → awardee Name (so we can check first_DoM_FY)
    raw_pn_to_name = dict(zip(raw_grants["Project Number"].astype(str), raw_grants["Name"]))

    rows = []
    for _, r in user_df.iterrows():
        pn = r["Project Number"]
        if pn in gen_proj or pn == "nan" or not pn:
            continue
        contact = str(r.get("Contact PI Name", r.get("Project PI", ""))).strip()
        org = str(r.get("Organization Name", r.get("Organization", ""))).strip()
        fy_raw = r.get("Fiscal Year")
        try:
            fy = int(fy_raw) if pd.notna(fy_raw) else None
        except (ValueError, TypeError):
            fy = None
        activity = str(r.get("Activity", "")).strip()

        # Categorize with refined logic
        awardee = raw_pn_to_name.get(pn)
        if pn in step3_proj and awardee:
            first_yr = first_dom.get(awardee)
            if first_yr and fy and fy < first_yr:
                reason = f"Pre-DoM-year — correctly excluded (PI={awardee}, grant FY{fy} < first DoM AY{first_yr})"
            else:
                reason = f"Excluded by GENEROUS filter (PI={awardee}, FY{fy}, first DoM AY{first_yr})"
        elif pn in raw_proj:
            reason = f"Removed by Step 3 same-name-collision rule (PI={awardee})"
        elif fy and (fy < 2021 or fy > 2026):
            reason = f"Outside our query window (FY{fy})"
        else:
            contact_match = None
            for n in awardee_names:
                parts = n.split()
                if len(parts) >= 2 and parts[-1].upper() in contact.upper():
                    contact_match = n
                    break
            if contact_match:
                reason = f"API didn't return this grant for {contact_match} (name-form mismatch?)"
            else:
                reason = "Contact PI not in DoM awardee list (user export may include extra names)"

        rows.append({
            "Project Number": pn,
            "Activity": activity,
            "Contact PI": contact,
            "Organization": org,
            "Fiscal Year": fy,
            "Likely Reason Missing": reason,
            "Title (truncated)": str(r.get("Project Title", ""))[:100],
        })
    return pd.DataFrame(rows)


# ── STEP 5: BY-AWARD-TYPE AGGREGATES ────────────────────────────────────────
def by_award_type(dom: pd.DataFrame, summary: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for award_type in ["Pilot", "Junior"]:
        recipients = dom[dom["Award"].str.startswith(award_type)]["Name"].unique()
        sub = summary[summary["Name"].isin(recipients)]
        dom_total = sub[f"{award_type} $"].sum()
        nih_dir_gen = sub["NIH Direct (Generous)"].sum()
        nih_dir_cons = sub["NIH Direct (Conservative)"].sum()
        nih_ind_gen = sub["NIH Indirect (Generous)"].sum()
        nih_ind_cons = sub["NIH Indirect (Conservative)"].sum()
        rows.append({
            "Award Type": award_type,
            "Recipients": len(recipients),
            "DoM $ (this type only)": dom_total,
            "Recipients with any post-DoM NIH funding (gen)": (sub["NIH Direct (Generous)"] > 0).sum(),
            "NIH Direct (Generous)": nih_dir_gen,
            "NIH Indirect (Generous)": nih_ind_gen,
            "ROI Generous": round(nih_dir_gen / dom_total, 1) if dom_total else 0,
            "NIH Direct (Conservative)": nih_dir_cons,
            "NIH Indirect (Conservative)": nih_ind_cons,
            "ROI Conservative": round(nih_dir_cons / dom_total, 1) if dom_total else 0,
        })
    return pd.DataFrame(rows)


# ── STEP 6: CROSS-REFERENCE WITH USER'S MANUAL EXPORTS ──────────────────────
def cross_reference(grants: pd.DataFrame) -> pd.DataFrame:
    issues = []
    # Compare against the Pilot+JrFac export (the appropriate one for this analysis)
    for fname, label in [(USER_EXPORT_SUB, "Pilot+JrFac export (primary)"),
                         (USER_EXPORT_ALL, "All-types export (incl. GT97; for reference only)")]:
        try:
            user_df = pd.read_excel(fname, sheet_name="Export Sheet")
            user_proj = set(user_df["Project Number"].dropna().astype(str).str.strip())
            our_proj = set(grants["Project Number"].dropna().astype(str).str.strip())

            missing_in_ours = user_proj - our_proj
            missing_in_user = our_proj - user_proj

            issues.append({
                "Source": label,
                "User export grant count": len(user_proj),
                "Our pipeline grant count (filtered)": len(our_proj),
                "In user export but NOT in ours": len(missing_in_ours),
                "In ours but NOT in user export": len(missing_in_user),
                "Sample missing-in-ours": ", ".join(list(missing_in_ours)[:5]),
                "Sample missing-in-user": ", ".join(list(missing_in_user)[:5]),
            })
        except Exception as e:
            issues.append({"Source": label, "Error": str(e)})
    return pd.DataFrame(issues)


# ── STEP 7: BUILD EXCEL ─────────────────────────────────────────────────────
def build_excel(summary, by_type, bu_grants, ext_grants, xref, narrative, missing):
    wb = Workbook()

    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    data_font = Font(name="Arial", size=11)
    money_fmt = '"$"#,##0'
    ratio_fmt = '0.0"x"'
    thin = Border(bottom=Side(style="thin", color="D9D9D9"))
    alt = PatternFill("solid", fgColor="F2F7FB")
    warn = PatternFill("solid", fgColor="FFF3CD")

    def write(ws, headers, data, money_cols=None, ratio_cols=None):
        money_cols = set(money_cols or [])
        ratio_cols = set(ratio_cols or [])
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for i, (_, row) in enumerate(data.iterrows(), 2):
            for col, h in enumerate(headers, 1):
                v = row.get(h, "")
                c = ws.cell(row=i, column=col, value=v if pd.notna(v) else "")
                c.font = data_font
                c.border = thin
                if i % 2 == 0:
                    c.fill = alt
                if h in money_cols:
                    c.number_format = money_fmt
                    c.alignment = Alignment(horizontal="right")
                if h in ratio_cols:
                    c.number_format = ratio_fmt
                    c.alignment = Alignment(horizontal="right")
        if len(data):
            ws.auto_filter.ref = f"A1:{chr(64+len(headers))}{len(data)+1}"
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30

    # Tab 1: Summary
    ws1 = wb.active
    ws1.title = "Investigator Summary"
    h1 = [
        "Name", "Section", "FirstDoMFY", "LastDoMFY", "DoMYears", "DoMAwards",
        "Pilot $", "Junior $", "DoMTotal",
        "NIH Direct (Generous)", "NIH Indirect (Generous)", "Grant Count (Generous)",
        "ROI Generous (Direct/DoM)",
        "NIH Direct (Conservative)", "NIH Indirect (Conservative)", "Grant Count (Conservative)",
        "ROI Conservative (Direct/DoM)",
    ]
    money_cols = ["Pilot $", "Junior $", "DoMTotal",
                  "NIH Direct (Generous)", "NIH Indirect (Generous)",
                  "NIH Direct (Conservative)", "NIH Indirect (Conservative)"]
    ratio_cols = ["ROI Generous (Direct/DoM)", "ROI Conservative (Direct/DoM)"]
    write(ws1, h1, summary, money_cols=money_cols, ratio_cols=ratio_cols)
    widths = [28, 26, 10, 10, 22, 22, 12, 12, 14, 18, 18, 14, 16, 18, 18, 14, 16]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[chr(64+i)].width = w

    # Tab 2: By Award Type
    ws2 = wb.create_sheet("By Award Type")
    h2 = list(by_type.columns)
    money_t = [c for c in h2 if "$" in c or "NIH" in c]
    ratio_t = [c for c in h2 if c.startswith("ROI")]
    write(ws2, h2, by_type, money_cols=money_t, ratio_cols=ratio_t)
    for i in range(1, len(h2)+1):
        ws2.column_dimensions[chr(64+i)].width = 20

    # Tab 3 & 4: BU/BMC and External grant detail
    grant_h = [
        "Name", "Project Number", "Activity Code", "Title", "Organization",
        "Fiscal Year", "Direct Cost", "Indirect Cost", "Award Amount",
        "IC", "Contact PI", "Start Date", "End Date", "Is Active",
    ]
    ws3 = wb.create_sheet("NIH Grants - BU_BMC")
    write(ws3, grant_h, bu_grants, money_cols=["Direct Cost", "Indirect Cost", "Award Amount"])
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["D"].width = 50
    ws3.column_dimensions["E"].width = 35

    ws4 = wb.create_sheet("NIH Grants - External")
    write(ws4, grant_h, ext_grants, money_cols=["Direct Cost", "Indirect Cost", "Award Amount"])
    ws4.column_dimensions["A"].width = 28
    ws4.column_dimensions["B"].width = 22
    ws4.column_dimensions["D"].width = 50
    ws4.column_dimensions["E"].width = 35

    # Tab 5: Per-Investigator Narrative
    ws5 = wb.create_sheet("Investigator Narrative")
    h5 = list(narrative.columns)
    money_n = [c for c in h5 if "DoM Total" in c or "NIH Direct" in c]
    ratio_n = [c for c in h5 if c.startswith("ROI")]
    write(ws5, h5, narrative, money_cols=money_n, ratio_cols=ratio_n)
    widths_n = {1: 26, 2: 22, 3: 40, 4: 14, 5: 18, 6: 14, 7: 70, 8: 18, 9: 14, 10: 70}
    for i, w in widths_n.items():
        ws5.column_dimensions[chr(64+i)].width = w
    # Wrap text in narrative columns
    for r in range(2, len(narrative)+2):
        ws5.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        ws5.cell(row=r, column=7).alignment = Alignment(wrap_text=True, vertical="top")
        ws5.cell(row=r, column=10).alignment = Alignment(wrap_text=True, vertical="top")
        ws5.row_dimensions[r].height = 60

    # Tab 6: Cross-reference summary
    ws6 = wb.create_sheet("Cross-Reference")
    h6 = list(xref.columns)
    write(ws6, h6, xref)
    for row_idx in range(2, len(xref)+2):
        for col_idx in range(1, len(h6)+1):
            ws6.cell(row=row_idx, column=col_idx).fill = warn
    for i in range(1, len(h6)+1):
        ws6.column_dimensions[chr(64+i)].width = 25

    # Tab 7: Missing-grants investigation
    ws7 = wb.create_sheet("Missing-Grants Investigation")
    h7 = list(missing.columns)
    write(ws7, h7, missing)
    widths_m = {1: 22, 2: 10, 3: 26, 4: 32, 5: 10, 6: 55, 7: 70}
    for i, w in widths_m.items():
        if i <= len(h7):
            ws7.column_dimensions[chr(64+i)].width = w

    wb.save(OUT_FILE)


# ── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    print("Step 1: Loading DoM Award Tracker (Pilot + Junior only — GT97 excluded)...")
    dom = load_dom_awards()
    print(f"  {len(dom)} award rows, {dom['Name'].nunique()} unique investigators")
    print(f"  Award types: {dom['Award'].value_counts().to_dict()}")
    print(f"  Total DoM $: ${dom['Amount'].sum():,.0f}")
    print(f"  By type: {dom.groupby('Award')['Amount'].sum().to_dict()}\n")

    names = sorted(dom["Name"].unique())

    if CACHE_FILE.exists():
        print(f"Step 2: Loading cached NIH RePORTER results from {CACHE_FILE.name}...")
        raw_grants = pd.read_csv(CACHE_FILE)
        # Filter cache to current name list — avoids leaking GT97-only investigators
        before = len(raw_grants)
        raw_grants = raw_grants[raw_grants["Name"].isin(set(names))].copy()
        dropped = before - len(raw_grants)
        print(f"  Cached: {len(raw_grants)} grant-year records (filtered out {dropped} not in current name list; delete cache to re-query)\n")
    else:
        print(f"Step 2: Querying NIH RePORTER for {len(names)} investigators (FY2021-2026)...")
        raw_grants = query_nih(names)
        raw_grants.to_csv(CACHE_FILE, index=False)
        print(f"  Raw: {len(raw_grants)} grant-year records (cached to {CACHE_FILE.name})\n")

    print("Step 3: Filtering false positives...")
    grants = filter_grants(raw_grants)
    print(f"  Filtered: {len(grants)} grant-year records, {grants['Name'].nunique()} investigators\n")

    print("Step 4: Computing ROI under both definitions...")
    summary, gen_grants, cons_grants = compute_roi(dom, grants)
    print(f"  Investigators with any generous post-DoM NIH funding: {(summary['NIH Direct (Generous)']>0).sum()}/{len(summary)}")
    print(f"  Investigators with any conservative post-DoM NIH funding: {(summary['NIH Direct (Conservative)']>0).sum()}/{len(summary)}\n")

    print("Step 5: By-award-type aggregates...")
    by_type = by_award_type(dom, summary)
    print(by_type[["Award Type", "Recipients", "DoM $ (this type only)",
                   "ROI Generous", "ROI Conservative"]].to_string(index=False))
    print()

    print("Step 6: Cross-referencing with user's manual exports...")
    xref = cross_reference(gen_grants)
    print(xref[["Source", "User export grant count", "Our pipeline grant count (filtered)",
                "In user export but NOT in ours", "In ours but NOT in user export"]].to_string(index=False))
    print()

    # Build grant detail tabs from generous filter
    bu_grants = gen_grants[gen_grants["Location"] == "BU/BMC"].sort_values(["Name", "Fiscal Year"]).reset_index(drop=True)
    ext_grants = gen_grants[gen_grants["Location"] == "External"].sort_values(["Name", "Fiscal Year"]).reset_index(drop=True)

    print("Step 7: Building per-investigator narrative...")
    narrative = build_narrative(dom, gen_grants, cons_grants)
    print(f"  {len(narrative)} investigator rows\n")

    print("Step 8: Investigating grants in user export but not in our pipeline...")
    missing = investigate_missing(gen_grants, grants, raw_grants, dom)
    if len(missing):
        print(f"  {len(missing)} grants flagged. Reason breakdown:")
        for reason, n in missing["Likely Reason Missing"].value_counts().items():
            print(f"    {n}: {reason}")
    print()

    print("Step 9: Building Excel...")
    build_excel(summary, by_type, bu_grants, ext_grants, xref, narrative, missing)
    print(f"  Saved to: {OUT_FILE}")

    # Headline numbers
    total_dom = dom["Amount"].sum()
    total_dir_gen = summary["NIH Direct (Generous)"].sum()
    total_ind_gen = summary["NIH Indirect (Generous)"].sum()
    total_dir_cons = summary["NIH Direct (Conservative)"].sum()
    total_ind_cons = summary["NIH Indirect (Conservative)"].sum()
    print(f"\n── HEADLINE NUMBERS ──")
    print(f"  DoM total invested (Pilot + Junior, AY22-AY26): ${total_dom:,.0f}")
    print(f"")
    print(f"  GENEROUS (active grants from first DoM year onward):")
    print(f"    NIH Direct:   ${total_dir_gen:,.0f}  →  ROI: {total_dir_gen/total_dom:.1f}x")
    print(f"    NIH Indirect: ${total_ind_gen:,.0f}")
    print(f"    NIH Total:    ${total_dir_gen+total_ind_gen:,.0f}  →  ROI: {(total_dir_gen+total_ind_gen)/total_dom:.1f}x")
    print(f"")
    print(f"  CONSERVATIVE (only grants started AT or AFTER first DoM year):")
    print(f"    NIH Direct:   ${total_dir_cons:,.0f}  →  ROI: {total_dir_cons/total_dom:.1f}x")
    print(f"    NIH Indirect: ${total_ind_cons:,.0f}")
    print(f"    NIH Total:    ${total_dir_cons+total_ind_cons:,.0f}  →  ROI: {(total_dir_cons+total_ind_cons)/total_dom:.1f}x")


if __name__ == "__main__":
    main()
